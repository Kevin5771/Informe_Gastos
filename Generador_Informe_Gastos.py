import openpyxl
import matplotlib.pyplot as plt
import os


#Presentación
def presentacion():
    print("------GENERADOR DE INFORME DE GASTOS------")

#Carga de Datos al Formato Xlsx
def cargar_datos_desde_excel():
    if os.path.exists("informe_gastos.xlsx"):
        workbook = openpyxl.load_workbook("informe_gastos.xlsx")
        sheet = workbook.active

        gastos = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            fecha, descripcion, monto, categoria = row
            gastos.append((fecha, descripcion, float(monto), categoria))

        return gastos

    return []

def guardar_en_excel(gastos):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Gastos"

    sheet['A1'] = "Fecha"
    sheet['B1'] = "Descripción"
    sheet['C1'] = "Monto"
    sheet['D1'] = "Categoría"

    for index, gasto in enumerate(gastos, start=2):
        sheet.cell(row=index, column=1, value=gasto[0])
        sheet.cell(row=index, column=2, value=gasto[1])
        sheet.cell(row=index, column=3, value=gasto[2])
        sheet.cell(row=index, column=4, value=gasto[3])

    workbook.save("informe_gastos.xlsx")

def ingresar_gastos():
    gastos = cargar_datos_desde_excel()

    while True:
        print("\n1. Agregar gasto")
        print("2. Editar gasto")
        print("3. Buscar gasto")
        print("4. Mostrar resumen")
        print("5. Generar gráfico")
        print("6. Salir")
        opcion = input("Seleccione una opción: ")

        if opcion == '1':
            fecha = input("Ingrese la fecha del gasto: ")
            descripcion = input("Ingrese la descripción del gasto: ")
            monto = float(input("Ingrese el monto del gasto: "))
            categoria = input("Ingrese la categoría del gasto: ")
            gastos.append((fecha, descripcion, monto, categoria))
        elif opcion == '2':
            indice = int(input("Ingrese el índice del gasto a editar: ")) - 1
            if 0 <= indice < len(gastos):
                fecha = input("Ingrese la nueva fecha del gasto: ")
                descripcion = input("Ingrese la nueva descripción del gasto: ")
                monto = float(input("Ingrese el nuevo monto del gasto: "))
                categoria = input("Ingrese la nueva categoría del gasto: ")
                gastos[indice] = (fecha, descripcion, monto, categoria)
                print("Gasto editado con éxito.")
            else:
                print("Índice de gasto no válido.")
        elif opcion == '3':
            busqueda = input("Ingrese una palabra clave para buscar: ")
            resultados = [gasto for gasto in gastos if busqueda.lower() in gasto[1].lower()]
            if resultados:
                print("\nResultados de la búsqueda:")
                for i, resultado in enumerate(resultados, start=1):
                    print(f"{i}. Fecha: {resultado[0]}, Descripción: {resultado[1]}, Monto: ${resultado[2]}, Categoría: {resultado[3]}")
            else:
                print("No se encontraron resultados para la búsqueda.")
        elif opcion == '4':
            resumen = calcular_resumen_gastos(gastos)
            if resumen:
                print("\nResumen de Gastos:")
                print(f"Número total de gastos: {resumen['numero_de_gastos']}")
                print(f"Gasto más caro: Fecha: {resumen['gasto_mas_caro'][0]}, Descripción: {resumen['gasto_mas_caro'][1]}, Monto: Q.{resumen['gasto_mas_caro'][2]}, Categoría: {resumen['gasto_mas_caro'][3]}")
                print(f"Gasto más barato: Fecha: {resumen['gasto_mas_barato'][0]}, Descripción: {resumen['gasto_mas_barato'][1]}, Monto: Q.{resumen['gasto_mas_barato'][2]}, Categoría: {resumen['gasto_mas_barato'][3]}")
                print(f"Monto total de gastos: Q.{resumen['monto_total_gastos']}")
            else:
                print("No se registraron gastos.")
        elif opcion == '5':
            generar_grafico_categorias(gastos)
        elif opcion == '6':
            guardar_en_excel(gastos)
            print("Los datos se han guardado en el archivo 'informe_gastos.xlsx'.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def calcular_resumen_gastos(gastos):
    if not gastos:
        return None

    total_gastos = sum(gasto[2] for gasto in gastos)
    gasto_mas_caro = max(gastos, key=lambda x: x[2])
    gasto_mas_barato = min(gastos, key=lambda x: x[2])

    return {
        "numero_de_gastos": len(gastos),
        "gasto_mas_caro": gasto_mas_caro,
        "gasto_mas_barato": gasto_mas_barato,
        "monto_total_gastos": total_gastos
    }

def generar_grafico_categorias(gastos):
    categorias = {}
    for gasto in gastos:
        categoria = gasto[3]
        if categoria in categorias:
            categorias[categoria] += gasto[2]
        else:
            categorias[categoria] = gasto[2]

    plt.figure(figsize=(8, 8))
    plt.pie(categorias.values(), labels=categorias.keys(), autopct='%1.1f%%')
    plt.title("Distribución de Gastos por Categoría")
    plt.show()

#Función Principal
presentacion()
ingresar_gastos()